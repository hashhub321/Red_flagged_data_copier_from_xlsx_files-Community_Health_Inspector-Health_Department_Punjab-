import streamlit as st
import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries
import io
import zipfile  # NEW IMPORT REQUIRED

# ... [Keep your CSS, UI setup, and helper functions the same] ...

st.set_page_config(page_title="Flagged Data Extractor", layout="centered")

st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600&display=swap');
    html, body, [class*="css"] { font-family: 'Inter', sans-serif !important; }
    [data-testid="stFileUploader"] { margin-bottom: 0.5rem !important; }
    [data-testid="stButton"], [data-testid="stDownloadButton"] { margin-top: 0.5rem !important; margin-bottom: 1rem !important; }
    [data-testid="stButton"] button, [data-testid="stDownloadButton"] button { font-size: 18px !important; font-weight: 600 !important; padding: 0.75rem !important; }
    [data-testid="stDownloadButton"] button { background-color: #16a34a !important; border-color: #16a34a !important; color: white !important; }
    [data-testid="stDownloadButton"] button:hover { background-color: #15803d !important; border-color: #15803d !important; }
    [data-testid="stModal"] { display: flex !important; align-items: center !important; justify-content: center !important; }
    [data-testid="stModal"] > div { margin: auto !important; }
    </style>
""", unsafe_allow_html=True)

st.title("Flagged Data Extractor")
st.write("Upload your data files (which have colored duplicate inputs) and submission file (where all the duplicate data will be saved). Processed results will be available for download.")

def find_first_empty_row(ws, key_columns):
    row = 2
    while True:
        if all(ws.cell(row=row, column=col).value in (None, '') for col in key_columns):
            return row
        row += 1
        if row > ws.max_row + 1000:
            return row

def clean_cnic(value):
    if value is None or value == '': return None
    if isinstance(value, str):
        stripped = value.strip()
        return int(stripped) if stripped.isdigit() else stripped
    if isinstance(value, float):
        return int(value) if value.is_integer() else value
    return value

# ------------------------------------------------------------------
# NEW HELPER FUNCTION — decides whether a fill's color is effectively
# "white" (i.e. visually indistinguishable from an unfilled cell).
# A solid white fill should NOT count as "flagged/colored", since it
# carries no visible highlight information.
# ------------------------------------------------------------------
def is_white_fill_color(fill):
    if not fill or fill.patternType in (None, 'none'):
        return True  # no pattern at all -> treat as "white/no fill"

    color = fill.fgColor
    if color is None:
        return True

    if color.type == 'rgb':
        # FFFFFFFF = opaque white, 00FFFFFF = fully transparent white
        return color.rgb in ('FFFFFFFF', '00FFFFFF', None)

    if color.type == 'theme':
        # Theme 0 is "Background 1" (white) in the default Office theme.
        # A tint can lighten/darken it, but tint 0 (or unset) is pure white.
        return color.theme == 0 and (color.tint in (0, None))

    if color.type == 'indexed':
        # 64 = "None"/system default, 9 = legacy palette white
        return color.indexed in (64, 9)

    return False

def is_cell_flagged(cell):
    if cell.value in (None, ''):
        return False

    # Condition 1: The cell background has any non-default, non-white pattern fill
    # (covers 'solid' and also less-common patterns like 'darkGrid',
    # 'lightHorizontal', etc., so exotic fills aren't missed — but a
    # plain white fill is treated the same as "no fill".)
    if cell.fill and cell.fill.patternType not in (None, 'none') and not is_white_fill_color(cell.fill):
        return True

    # Condition 2: The text has a custom color (ignoring default black/white/auto)
    if cell.font and cell.font.color:
        color = cell.font.color
        # Check RGB values (ignoring pure black and pure white)
        if color.type == 'rgb' and color.rgb not in ('FF000000', '00000000', 'FFFFFFFF', '00FFFFFF'):
            return True
        # Check Theme colors (Theme 0 and 1 are default black/white)
        if color.type == 'theme' and color.theme not in (0, 1):
            return True
        # Check Indexed colors (8 and 64 are default black/auto in legacy Excel)
        if color.type == 'indexed' and color.indexed not in (8, 64):
            return True

    return False

# ------------------------------------------------------------------
# NEW HELPER FUNCTION — this is the only new "logic" piece added.
# It wraps is_cell_flagged() and decides whether a cell should be
# treated as extractable, based on the mode chosen by the user.
#   mode = "colored"   -> keep old behaviour (extract colored cells)
#   mode = "uncolored" -> extract cells that have a value but are NOT colored
# ------------------------------------------------------------------
def should_extract_cell(cell, mode):
    if cell.value in (None, ''):
        return False
    flagged = is_cell_flagged(cell)
    if mode == "uncolored":
        return not flagged
    return flagged  # default / "colored" mode

@st.dialog("✅ Processing Complete!")
def show_download_popup(file_data):
    st.write("Your files have been merged. Tap the button below to save the result.")
    st.download_button(
        label="⬇️ Download Completed File",
        data=file_data,
        file_name="Flagged_Data_Extractor_Result.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True
    )

def write_cnic(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=clean_cnic(value))
    if isinstance(cell.value, int):
        cell.number_format = '0'

with st.form("upload_and_process_form", clear_on_submit=False):
    submission_file = st.file_uploader("1. Upload Submission File ", type=["xlsx"], key="sub_key")
    source_files = st.file_uploader("2. Upload Source Files (Colored-Data)", type=["xlsx"], accept_multiple_files=True, key="src_key")

    # ------------------------------------------------------------------
    # NEW UI ELEMENT — toggle to choose extraction target.
    # Stored as a plain string ("colored" / "uncolored") so it can be
    # passed straight into should_extract_cell().
    # ------------------------------------------------------------------
    extraction_choice = st.radio(
        "3. Which house codes should be extracted?",
        options=["Colored house codes (default)", "Un-colored house codes"],
        index=0,
        help="Choose 'Colored' if the flagged/duplicate house codes are the ones highlighted. "
             "Choose 'Un-colored' if instead the highlighted ones should be IGNORED, "
             "and the plain/un-highlighted house codes are the ones to extract."
    )
    extraction_mode = "colored" if extraction_choice.startswith("Colored") else "uncolored"

    submitted = st.form_submit_button("Process Files", type="primary", use_container_width=True)

if submitted:
    if not submission_file or not source_files:
        st.error("Please upload both the submission file and at least one source file.")
    else:
        try:
            # Check if submission file is locked/empty
            if submission_file.size == 0:
                st.error("❌ Submission file is 0 bytes. If it is open in another app, close it and try again.")
                st.stop()

            target_wb = openpyxl.load_workbook(submission_file)
            target_ws = target_wb.active
            target_headers = {str(cell.value).strip(): idx for idx, cell in enumerate(target_ws[1], 1) if cell.value}

            key_cols_for_emptiness = [
                target_headers[h] for h in
                ['District', 'Tehsil', 'UC-Name', 'Head of family CNIC', 'CHI CNIC', 'CHI Name', 'House code']
                if h in target_headers
            ]

            target_row = find_first_empty_row(target_ws, key_cols_for_emptiness)
            start_row = target_row
            base_mapping = {'District': 'District', 'Tehsil': 'Tehsil', 'UC': 'UC-Name', 'HeadOfFamilyCNIC': 'Head of family CNIC'}

            for uploaded_source in source_files:
                # Check for locked/empty source files
                if uploaded_source.size == 0:
                    st.error(f"❌ Cannot read '{uploaded_source.name}'. It may be open in another app or is a cloud shortcut. Please close it and re-upload.")
                    st.stop()

                try:
                    wb = openpyxl.load_workbook(uploaded_source, data_only=True)
                except zipfile.BadZipFile:
                    st.error(f"❌ '{uploaded_source.name}' is unreadable or locked by Android. Please make sure the file is closed on your phone before uploading.")
                    st.stop()

                ws = wb.active
                source_headers = {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1], 1) if cell.value}

                # ... [Keep your exact row processing logic here] ...
                for row in ws.iter_rows(min_row=2):
                    for i in range(1, 5):
                        cadre_col = f'Cadre {i}'
                        cnic_col = f'CNICofCadre {i}'
                        house_col = f'HouseCode {i}'

                        if house_col in source_headers:
                            house_cell_idx = source_headers[house_col] - 1
                            house_cell = row[house_cell_idx]

                            # CHANGED LINE: was `if is_cell_flagged(house_cell):`
                            # now routes through the mode-aware helper.
                            if should_extract_cell(house_cell, extraction_mode):
                                for src, tgt in base_mapping.items():
                                    if src in source_headers and tgt in target_headers:
                                        value = row[source_headers[src] - 1].value
                                        if src == 'HeadOfFamilyCNIC':
                                            write_cnic(target_ws, target_row, target_headers[tgt], value)
                                        else:
                                            target_ws.cell(row=target_row, column=target_headers[tgt], value=value)

                                if cadre_col in source_headers and 'CHI Name' in target_headers:
                                    target_ws.cell(row=target_row, column=target_headers['CHI Name'], value=row[source_headers[cadre_col] - 1].value)

                                if cnic_col in source_headers and 'CHI CNIC' in target_headers:
                                    write_cnic(target_ws, target_row, target_headers['CHI CNIC'], row[source_headers[cnic_col] - 1].value)

                                if 'House code' in target_headers:
                                    target_ws.cell(row=target_row, column=target_headers['House code'], value=house_cell.value)

                                target_row += 1

            last_written_row = target_row - 1
            for table in target_ws.tables.values():
                min_col, min_row, max_col, max_row = range_boundaries(table.ref)
                new_max_row = max(max_row, last_written_row)
                table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_max_row}"

            output = io.BytesIO()
            target_wb.save(output)
            output.seek(0)

            show_download_popup(output)

        except zipfile.BadZipFile:
             st.error("❌ The Submission file is unreadable or locked. Make sure it is closed on your phone.")
        except Exception as e:
            st.error(f"An error occurred: {e}")
